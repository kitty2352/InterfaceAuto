import xlwt
import xlrd
import openpyxl as xl
import pandas as pd
import os

class OperateExcel:

    def __init__(self, file_name=None, sheet_id=None):
        if file_name:
            self.file_name = file_name
            self.sheet_id = sheet_id
        else:
            self.file_name = '../data/interfaceData.xlsx'
            self.sheet_id = 0
        self.data = self.get_data()

    def write_excel_file(self, row, col, content):
        result_path = os.path.join(self.file_name)
        if os.path.exists(result_path):
            workbook = xl.load_workbook(result_path)
            sheet = workbook.active
            sheet.cell(row=row, column=col, value=content)
            workbook.save(result_path)


    # 获取excel中的数据
    def get_data(self):
        data = xlrd.open_workbook(self.file_name)
        tables = data.sheets()[self.sheet_id]
        return tables

    # 获取excel中的行数
    def get_rows(self):
        return self.get_data().nrows

    # 获取excel中的列数
    def get_cols(self):
        return self.data.ncols

    # 获取行内容
    def get_row_content(self, row):
        return self.data.row_values(row)

    # 获取列内容
    def get_col_content(self, col):
        return self.data.col_values(col)

    # 获取行数
    def get_line(self):
        return self.data.nrows

    # 获取某个单元格内容
    def get_cell_value(self, row, col):
        return self.data.cell(row, col).value


if __name__ == '__main__':
    excelObj = OperateExcel()
    excelObj.write_excel_file(5, 1, "test")
